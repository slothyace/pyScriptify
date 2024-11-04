# StdLib
import os
import shutil
import tkinter as tk
import pyFunctions.baseFunctions.modMan as modMan
import copy

# Required External Lib
extLib = ["pandas", "openpyxl", "openpyxl.utils"]
modMan.modMan(extLib)

import pandas
import openpyxl
import openpyxl.utils

def consoleClear(console):
    console.delete("1.0", tk.END)

def consolePrint(console, string):
    console.insert(tk.END, f"{string}\n\n")
    console.see(tk.END)
    console.update_idletasks()

def consoleSee(console):
    console.see(tk.END)
    console.update_idletasks()

def deleteFile(console, filePath):
    try:
        os.remove(filePath)
        console.insert(tk.END, f"{filePath} deleted.\n\n")

    except Exception as e:
        console.insert(tk.END, f"{filePath} couldn't be deleted. {e}\n\n")

    console.see(tk.END)
    console.update_idletasks()

def deleteFolder(console, folderPath):
    try:
        shutil.rmtree(folderPath)
        console.insert(tk.END, f"{folderPath} deleted.\n\n")

    except Exception as e:
        console.insert(tk.END, f"{folderPath} couldn't be deleted. {e}\n\n")

    console.see(tk.END)
    console.update_idletasks()

def csvToExcel(console, filePath):
    try:
        pathOut = filePath.replace(".csv", ".xlsx")
        dataFrame = pandas.read_csv(filePath)
        dataFrame.to_excel(pathOut, index=False)
        console.insert(tk.END, f"File {os.path.basename(filePath).split(f".csv")[0]} converted.\n\n")
    
    except Exception as e:
        console.insert(tk.END, f"File {os.path.basename(filePath).split(f".csv")[0]} couldn't be converted. {e}\n\n")

    console.see(tk.END)
    console.update_idletasks()

def excelToCsv(console, filePath):
    try:
        pathOut = filePath.replace(".xlsx", ".csv")
        dataFrame = pandas.read_excel(filePath, engine="openpyxl")
        dataFrame.to_csv(pathOut, index=False)
        console.insert(tk.END, f"File {os.path.basename(filePath).split(f".xlsx")[0]} converted.\n\n")

    except Exception as e:
        console.insert(tk.END, f"File {os.path.basename(filePath).split(f".xlsx")[0]} couldn't be converted. {e}\n\n")

    console.see(tk.END)
    console.update_idletasks()

def listFilesInFolder(console, folderPath, extension):
    folderFiles = os.listdir(folderPath)

    if extension == "*":
        fileList = folderFiles
        pathOut = os.path.join(folderPath, f"_folderIndex.txt")

    else:
        fileList = [file for file in folderFiles if file.endswith(f".{extension}")]
        pathOut = os.path.join(folderPath, f"_{extension}Index.txt")

    with open(pathOut, "w") as pathOutContent:
        for file in fileList:
            pathOutContent.write(file.split(f".{extension}")[0] + f"\n")

    console.insert(tk.END, f"List created and saved to {pathOut}.\n\n")
    console.see(tk.END)
    console.update_idletasks()

def splitExcel(console, filePath):
    workbook = openpyxl.load_workbook(filePath)
    oriFileName = os.path.splitext(os.path.basename(filePath))[0]
    oriDirectory = os.path.dirname(filePath)
    pathOut = os.path.join(oriDirectory, f"{oriFileName}_splitSheets")
    os.makedirs(pathOut, exist_ok=True)

    for sheetName in workbook.sheetnames:
        sheet = workbook[sheetName]
        newWorkbook = openpyxl.Workbook()
        newSheet = newWorkbook.active

        for row in sheet.iter_rows():
            for cell in row:
                if cell.row <= sheet.max_row and cell.column <= sheet.max_column :
                    newSheet.cell(row = cell.row, column=cell.column, value=cell.value)

        for row in range(1, sheet.max_row +1):
            newSheet.row_dimensions[row].height = sheet.row_dimensions[row].height

        for col in range(1, sheet.max_column +1):
            colLetter = openpyxl.utils.get_column_letter(col)
            newSheet.column_dimensions[colLetter].width = sheet.column_dimensions[colLetter].width

        for mergedRange in sheet.merged_cells.ranges:
            newSheet.merge_cells(str(mergedRange))

        for row in sheet.iter_rows():
            for cell in row:
                if cell.row <= sheet.max_row and cell.column <= sheet.max_column:
                    newCell = newSheet.cell(row = cell.row, column = cell.column)
                    if cell.has_style:
                        newCell.font = copy.copy(cell.font)
                        newCell.border = copy.copy(cell.border)
                        newCell.fill = copy.copy(cell.fill)
                        newCell.number_format = copy.copy(cell.number_format)
                        newCell.protection = copy.copy(cell.protection)
                        newCell.alignment = copy.copy(cell.alignment)

        sheetPathOut = os.path.join(pathOut, f"{sheetName}.xlsx")
        newWorkbook.save(sheetPathOut)
        newWorkbook.close()
        console.insert(tk.END, f"{sheetName}.xlsx created.\n\n")
        console.see(tk.END)
        console.update_idletasks()
    
    console.insert(tk.END, f"Sheets in {oriFileName} has been split and saved to {pathOut}.\n\n")
    console.see(tk.END)
    console.update_idletasks()

def mergeExcel(console, folderPath):
    mergedWorkbook = openpyxl.Workbook()
    mergedWorkbook.remove(mergedWorkbook.active)
    oriDirectory = os.path.dirname(folderPath)

    for fileName in os.listdir(folderPath):
        if fileName.endswith(".xlsx"):
            filePath = os.path.join(folderPath, fileName)
            workbook = openpyxl.load_workbook(filePath)
            sheet = workbook.active
            newSheet = mergedWorkbook.create_sheet(title=fileName.split(f".xlsx")[0])

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.row <= sheet.max_row and cell.column <= sheet.max_column:
                        newSheet.cell(row = cell.row, column=cell.column, value=cell.value)

            for row in range(1, sheet.max_row +1):
                newSheet.row_dimensions[row].height = sheet.row_dimensions[row].height

            for col in range(1, sheet.max_column +1):
                colLetter = openpyxl.utils.get_column_letter(col)
                newSheet.column_dimensions[colLetter].width = sheet.column_dimensions[colLetter].width
            
            for mergedRange in sheet.merged_cells.ranges:
                newSheet.merge_cells(str(mergedRange))

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.row <= sheet.max_row and cell.column <= sheet.max_column:
                        newCell = newSheet.cell(row = cell.row, column = cell.column)
                        if cell.has_style:
                            newCell.font = copy.copy(cell.font)
                            newCell.border = copy.copy(cell.border)
                            newCell.fill = copy.copy(cell.fill)
                            newCell.number_format = copy.copy(cell.number_format)
                            newCell.protection = copy.copy(cell.protection)
                            newCell.alignment = copy.copy(cell.alignment)
            
            console.insert(tk.END, f"{fileName.split(f".xlsx")[0]} merged in.\n\n")
            console.see(tk.END)
            console.update_idletasks()

    pathOut = os.path.join(oriDirectory, os.path.basename(folderPath) + f"_mergedExcel.xlsx")
    mergedWorkbook.save(pathOut)
    console.insert(tk.END, f"Sheets in {os.path.basename(folderPath)} has been split and saved to {pathOut}.\n\n")
    console.see(tk.END)
    console.update_idletasks()
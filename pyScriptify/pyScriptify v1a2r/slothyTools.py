# PyScript Version
version = 'v1a2r-PyScript'

# Python Standard Libraries
import os
import sys
import subprocess
import csv
import glob
import tkinter as tk
from tkinter import filedialog, messagebox
import time

# External Libraries
print('Checking for external libraries')
installables = ["ttkbootstrap", "openpyxl", "pyfiglet"]
for library in installables:
    try:
        __import__(library)
    except ImportError:
        print(f'\n{library} is missing.')
        print('Attempting to install.')
        os.system(f'pip install {library}')
print('All required libraries present')

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import openpyxl
import pyfiglet

if not sys.executable.endswith('pythonw.exe'):
    try:
        sT = "slothyTools"
        sTart = pyfiglet.figlet_format(sT)
        os.system('cls')
        print(sTart, f"\nRelaunching slothyTools-{version} with pythonw")
        time.sleep(0.5)
        os.system('start pythonw.exe slothyTools.py')
        sys.exit()
    except Exception as e:
        log_error(f"Failed to relaunch with pythonw: {e}")
else:
    def delFile(file, console):
        try:
            os.remove(file)
            console.insert(tk.END, f'{file} deleted.\n\n')
        except:
            console.insert(tk.END, f'{file} couldn\'t be deleted.\n\n')
        
    def delFileFile(file, console):
        delChoice = messagebox.askyesno('Delete file', f'Delete original file?')
        if delChoice:
            delFile(file, console)
        else:
            console.insert(tk.END, f'Original file not deleted.\n\n')
        console.see(tk.END)
        console.update_idletasks()
        
    def delFileFolder(files, console):
        delChoice = messagebox.askyesno('Delete file', f'Delete original files?')
        if delChoice:
            for file in files:
                delFile(file, console)
        else:
            console.insert(tk.END, f'Original files not deleted.\n\n')
        console.see(tk.END)
        console.update_idletasks()
        
    def c2x(filePath, console):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(filePath, 'r', encoding='utf-8') as csvFile:
                reader = csv.reader(csvFile)
                for row in reader:
                    ws.append(row)
            xlsxFile = filePath.replace('.csv', '.xlsx')
            wb.save(xlsxFile)
            console.insert(tk.END, f'{filePath} converted.\n\n')
        except Exception as e:
            console.insert(tk.END, f'{filePath} couldn\'t be converted: {e}\n\n')
        console.see(tk.END)
        console.update_idletasks()
        
    def x2c(filePath, console):
        try:
            wb = openpyxl.load_workbook(filePath)
            ws = wb.active
            csvFile = filePath.replace('.xlsx', '.csv')
            with open(csvFile, 'w', newline='', encoding='utf-8') as xlsxFile:
                writer = csv.writer(xlsxFile)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            console.insert(tk.END, f'{filePath} converted.\n\n')
        except Exception as e:
            console.insert(tk.END, f'{filePath} couldn\'t be converted: {e}\n\n')
        console.see(tk.END)
        console.update_idletasks()
        
    def c2xFolder(folderPath, console):
        csvFiles = glob.glob(os.path.join(folderPath, '*.csv'))
        if not csvFiles:
            console.insert(tk.END, f'No .csv files found in {folderPath}\n\n')
            console.see(tk.END)
            console.update_idletasks()
            return
        convertedFiles = []
        for csvFile in csvFiles:
            c2x(csvFile, console)
            convertedFiles.append(csvFile)
        delFileFolder(convertedFiles, console)
        
    def x2cFolder(folderPath, console):
        xlsxFiles = glob.glob(os.path.join(folderPath, '*.xlsx'))
        if not xlsxFiles:
            console.insert(tk.END, f'No .xlsx files found in {folderPath}\n\n')
            console.see(tk.END)
            console.update_idletasks()
            return
        convertedFiles = []
        for xlsxFile in xlsxFiles:
            x2c(xlsxFile, console)
            convertedFiles.append(xlsxFile)
        delFileFolder(convertedFiles, console)
        
    def browse(convType, scale, console):
        if convType == 'c2x':
            if scale == 'file':
                filePath = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
                if filePath:
                    clrConsole()
                    c2x(filePath, console)
                    delFileFile(filePath, console)
                    
            elif scale == 'folder':
                folderPath = filedialog.askdirectory()
                if folderPath:
                    clrConsole()
                    c2xFolder(folderPath, console)
                    
        elif convType == 'x2c':
            if scale == 'file':
                filePath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
                if filePath:
                    clrConsole()
                    x2c(filePath, console)
                    delFileFile(filePath, console)
                    
            elif scale == 'folder':
                folderPath = filedialog.askdirectory()
                if folderPath:
                    clrConsole()
                    x2cFolder(folderPath, console)

    def clrConsole():
        console.delete('1.0', tk.END)

    winy = 550
    winx = 1.5 * winy

    root = ttk.Window(themename='darkly')
    root.title('slothyTools')
    root.geometry(f'{int(winx)}x{winy}')
    root.resizable(True, True)
    root.minsize(int(winx), winy)

    sToolsFrame = ttk.LabelFrame(root, text=' Welcome To ', bootstyle='success')
    sToolsFrame.pack(pady=(0,10), padx=10, fill='x')
    ttk.Label(sToolsFrame, text=pyfiglet.figlet_format('slothyTools'), font=('consolas', 8)).pack()
    ttk.Label(sToolsFrame, text=(f'Version: {version}'), font=('consolas', 8)).pack(side='right', padx=10)

    ttk.Separator(root, bootstyle='success', orient='horizontal').pack(fill='x')

    menuBtnFrame = ttk.Frame(root)
    menuBtnFrame.pack(pady=8, padx=8, fill='none')

    c2xButton = ttk.Menubutton(menuBtnFrame, text='CSV -> XLSX', bootstyle='success', width=15)
    c2xMenu = ttk.Menu(c2xButton)
    c2xButton['menu'] = c2xMenu

    c2xMenu.add_command(label='File', command=lambda: browse('c2x', 'file', console))
    c2xMenu.add_command(label='Folder', command=lambda: browse('c2x', 'folder', console))

    x2cButton = ttk.Menubutton(menuBtnFrame, text='XLSX -> CSV', bootstyle='success', width=15)
    x2cMenu = ttk.Menu(x2cButton)
    x2cButton['menu'] = x2cMenu

    x2cMenu.add_command(label='File', command=lambda: browse('x2c', 'file', console))
    x2cMenu.add_command(label='Folder', command=lambda: browse('x2c', 'folder', console))

    c2xButton.grid(row=0, column=0, padx=2, pady=2)
    x2cButton.grid(row=0, column=1, padx=2, pady=2)

    ttk.Separator(root, bootstyle='success', orient='horizontal').pack(fill='x')

    consoleFrame = ttk.LabelFrame(root, text=' Console ', bootstyle='success')
    consoleFrame.pack(pady=(0,10), padx=10, fill='both', expand=True)

    console = ttk.Text(consoleFrame, wrap=tk.WORD, width=1, height=1, font='consolas')
    console.pack(pady=0, padx=10, fill=tk.BOTH, expand=True)
    console.bind("<Key>", lambda e: 'break')
    console.insert(tk.END, 'Welcome to slothyTools\n\n')
    ttk.Button(consoleFrame, text='Clear Console', command=clrConsole, bootstyle='success', width=20).pack(pady=(0,10), padx=10, fill='x', expand=False)

    root.mainloop()

# Compiler Ready Version
version = 'v1a3'

# Python Standard Libraries
import os
import sys
import subprocess
import csv
import glob
import tkinter as tk
from tkinter import filedialog, messagebox
import time
import shutil

# External Libraries
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import openpyxl
import openpyxl.utils
import pyfiglet
import pyfiglet.fonts

def delFile(file, console):
    try:
        os.remove(file)
        console.insert(tk.END, f'{file} deleted\n\n')
    except:
        console.insert(tk.END, f'{file} couldn\'t be deleted\n\n')
    
def delFileFile(file, console):
    delChoice = messagebox.askyesno('Delete file', f'Delete original file?')
    if delChoice:
        delFile(file, console)
    else:
        console.insert(tk.END, f'Original file not deleted\n\n')
    console.see(tk.END)
    console.update_idletasks()
    
def delFileFolder(files, console):
    delChoice = messagebox.askyesno('Delete file', f'Delete original files?')
    if delChoice:
        for file in files:
            delFile(file, console)
    else:
        console.insert(tk.END, f'Original files not deleted\n\n')
    console.see(tk.END)
    console.update_idletasks()

def delFolder(folder, console) :
    try:
        shutil.rmtree(folder)
        console.insert(tk.END, f'{folder} deleted\n\n')
    except:
        console.insert(tk.END, f'{folder} couldn\'t be deleted\n\n')

def delFolderSingle(folder, console):
    delChoice = messagebox.askyesno('Delete folder', f'Delete original folder?')
    if delChoice:
        delFolder(folder, console)
    else:
        console.insert(tk.END, f'Original folder not deleted\n\n')
    console.see(tk.END)
    console.update_idletasks()

def delFolderMulti(folders, console):
    delChoice = messagebox.askyesno('Delete folders', f'Delete original folders?')
    if delChoice:
        for folder in folders:
            delFolder(folder, console)
    else:
        console.insert(tk.END, f'Original folders not deleted\n\n')
    console.see(tk.END)
    console.update_idletasks()

def c2x(filePath, console):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(filePath, 'r', encoding='utf-8') as csvFile:
            reader = csv.reader(csvFile)
            for row in reader:
                ws.append(row)
        xlsxFile = filePath.replace('.csv', '.xlsx')
        wb.save(xlsxFile)
        console.insert(tk.END, f'{filePath} converted\n\n')
    except Exception as e:
        console.insert(tk.END, f'{filePath} couldn\'t be converted: {e}\n\n')
    console.see(tk.END)
    console.update_idletasks()
    
def x2c(filePath, console):
    try:
        wb = openpyxl.load_workbook(filePath)
        ws = wb.active
        csvFile = filePath.replace('.xlsx', '.csv')
        with open(csvFile, 'w', newline='', encoding='utf-8') as xlsxFile:
            writer = csv.writer(xlsxFile)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        console.insert(tk.END, f'{filePath} converted\n\n')
    except Exception as e:
        console.insert(tk.END, f'{filePath} couldn\'t be converted: {e}\n\n')
    console.see(tk.END)
    console.update_idletasks()
    
def c2xFolder(folderPath, console):
    csvFiles = glob.glob(os.path.join(folderPath, '*.csv'))
    if not csvFiles:
        console.insert(tk.END, f'No .csv files found in {folderPath}\n\n')
        console.see(tk.END)
        console.update_idletasks()
        return
    convertedFiles = []
    for csvFile in csvFiles:
        c2x(csvFile, console)
        convertedFiles.append(csvFile)
    delFileFolder(convertedFiles, console)
    
def x2cFolder(folderPath, console):
    xlsxFiles = glob.glob(os.path.join(folderPath, '*.xlsx'))
    if not xlsxFiles:
        console.insert(tk.END, f'No .xlsx files found in {folderPath}\n\n')
        console.see(tk.END)
        console.update_idletasks()
        return
    convertedFiles = []
    for xlsxFile in xlsxFiles:
        x2c(xlsxFile, console)
        convertedFiles.append(xlsxFile)
    delFileFolder(convertedFiles, console)

def clrConsole():
    console.delete('1.0', tk.END)

def asunder(filePath, console):
    workbook = openpyxl.load_workbook(filePath)
    oriFileName = os.path.splitext(os.path.basename(filePath))[0]
    oriDir = os.path.dirname(filePath)
    saveFolder = os.path.join(oriDir, f"{oriFileName} sheets")
    os.makedirs(saveFolder, exist_ok=True)
    
    for sheetName in workbook.sheetnames:
        sheet = workbook[sheetName]
        nWorkbook = openpyxl.Workbook()
        newSheet = nWorkbook.active
        newSheet.title = sheetName
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row <= sheet.max_row and cell.column <= sheet.max_column:
                    newSheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    newSheet.row_dimensions[cell.row].height = sheet.row_dimensions[cell.row].height
                    newSheet.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width = sheet.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width
        
        for mergedRange in sheet.merged_cells.ranges:
            newSheet.merge_cells(str(mergedRange))
        
        newFilePath = os.path.join(saveFolder, f"{sheetName}.xlsx")
        nWorkbook.save(newFilePath)
        nWorkbook.close()
        console.insert(tk.END, f'{sheetName}.xlsx created\n\n')

    console.insert(tk.END, f'Sheets have been split into individual files in {saveFolder}\n\n')
        
def asunderFolder(folderPath, console):
    xlsxFiles = glob.glob(os.path.join(folderPath, '*.xlsx'))
    if not xlsxFiles:
        console.insert(tk.END, f'No .xlsx files found in {folderPath}\n\n')
        console.see(tk.END)
        console.update_idletasks()
        return

    convertedFiles = []
    for xlsxFile in xlsxFiles:
        asunder(xlsxFile, console)
        convertedFiles.append(xlsxFile)
    delFileFolder(convertedFiles, console)

def unasunder(folderPath, console):
    mergedWorkbook = openpyxl.Workbook()
    mergedWorkbook.remove(mergedWorkbook.active)
    oriDir = os.path.dirname(folderPath)

    for fileName in os.listdir(folderPath):
        if fileName.endswith('.xlsx'):
            filePath = os.path.join(folderPath, fileName)
            workbook = openpyxl.load_workbook(filePath)
            sheet = workbook.active
            newSheet = mergedWorkbook.create_sheet(title=sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    newSheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    newSheet.row_dimensions[cell.row].height = sheet.row_dimensions[cell.row].height
                    newSheet.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width = sheet.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width

            for mergedRange in sheet.merged_cells.ranges:
                newSheet.merge_cells(str(mergedRange))

        console.insert(tk.END, f'Merged {fileName}\n\n')

    outputFile = os.path.join(oriDir, os.path.basename(folderPath) + '_merged.xlsx')
    mergedWorkbook.save(outputFile)
    console.insert(tk.END, f'All sheets have been merged into {outputFile}\n\n')

def unasunderFolder(foldersPath, console):
    foldersUnasundered = []
    for folder in os.listdir(foldersPath):
        folderPath = os.path.join(foldersPath, folder)

        if os.path.isdir(folderPath):
            unasunder(folderPath, console)
            foldersUnasundered.append(folderPath)
    
    delFolderMulti(foldersUnasundered, console)

def insrtConsole(string, console):
    clrConsole()
    console.insert(tk.END, string)
    console.see(tk.END)
    console.update_idletasks()

def browse(convType, scale, console):
    if convType == 'c2x':
        if scale == 'file':
            filePath = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
            if filePath:
                clrConsole()
                c2x(filePath, console)
                delFileFile(filePath, console)
                
        elif scale == 'folder':
            folderPath = filedialog.askdirectory()
            if folderPath:
                clrConsole()
                c2xFolder(folderPath, console)
                
    elif convType == 'x2c':
        if scale == 'file':
            filePath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if filePath:
                clrConsole()
                x2c(filePath, console)
                delFileFile(filePath, console)
                
        elif scale == 'folder':
            folderPath = filedialog.askdirectory()
            if folderPath:
                clrConsole()
                x2cFolder(folderPath, console)
                
    elif convType == 'asunder':
        if scale == 'file':
            filePath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if filePath:
                clrConsole()
                asunder(filePath, console)
                delFileFile(filePath, console)
                
        elif scale == 'folder':
            folderPath = filedialog.askdirectory()
            if folderPath:
                clrConsole()
                asunderFolder(folderPath, console)

    elif convType == 'unasunder':
        if scale == 'single':
            folderPath = filedialog.askdirectory()
            if folderPath:
                clrConsole()
                unasunder(folderPath, console)
                delFolderSingle(folderPath, console)

        elif scale == 'multi':
            foldersPath = filedialog.askdirectory()
            if foldersPath:
                clrConsole()
                unasunderFolder(foldersPath, console)

winy = 550
winx = 1.5 * winy

root = ttk.Window(themename='darkly')
root.title('slothyTools')
root.geometry(f'{int(winx)}x{winy}')
root.resizable(True, True)
root.minsize(int(winx), winy)

sToolsFrame = ttk.LabelFrame(root, text=' Welcome To ', bootstyle='success')
sToolsFrame.pack(pady=(0,10), padx=10, fill='x')
ttk.Label(sToolsFrame, text=pyfiglet.figlet_format('slothyTools'), font=('consolas', 8)).pack()
ttk.Label(sToolsFrame, text=(f'Version: {version}'), font=('consolas', 8)).pack(side='right', padx=10)

ttk.Separator(root, bootstyle='success', orient='horizontal').pack(fill='x')

menuBtnFrame = ttk.Frame(root)
menuBtnFrame.pack(pady=8, padx=8, fill='none')

c2xButton = ttk.Menubutton(menuBtnFrame, text='CSV -> XLSX', bootstyle='success', width=15)
c2xMenu = ttk.Menu(c2xButton)
c2xButton['menu'] = c2xMenu

c2xMenu.add_command(label='File', command=lambda: browse('c2x', 'file', console))
c2xMenu.add_command(label='Folder', command=lambda: browse('c2x', 'folder', console))
c2xMenu.add_command(label='Info', command=lambda: insrtConsole(f'Converts CSV files into Excel files\n\n', console))

x2cButton = ttk.Menubutton(menuBtnFrame, text='XLSX -> CSV', bootstyle='success', width=15)
x2cMenu = ttk.Menu(x2cButton)
x2cButton['menu'] = x2cMenu

x2cMenu.add_command(label='File', command=lambda: browse('x2c', 'file', console))
x2cMenu.add_command(label='Folder', command=lambda: browse('x2c', 'folder', console))
x2cMenu.add_command(label='Info', command=lambda: insrtConsole(f'Converts Excel files into CSV files\n\n', console))

asunderButton = ttk.Menubutton(menuBtnFrame, text='Split XLSX sheets', bootstyle='success', width=15)
asunderMenu = ttk.Menu(asunderButton)
asunderButton['menu'] = asunderMenu

asunderMenu.add_command(label='File', command=lambda: browse('asunder', 'file', console))
asunderMenu.add_command(label='Folder', command=lambda: browse('asunder', 'folder', console))
asunderMenu.add_command(label='Info', command=lambda: insrtConsole(f'Splits the sheets in a Excel file into individual Excel files\n\n', console))

unasunderButton = ttk.Menubutton(menuBtnFrame, text='Merge XLSX files', bootstyle='success', width=15)
unasunderMenu = ttk.Menu(unasunderButton)
unasunderButton['menu'] = unasunderMenu

unasunderMenu.add_command(label='Single', command=lambda: browse('unasunder', 'single', console))
unasunderMenu.add_command(label='Multi', command=lambda: browse('unasunder', 'multi', console))
unasunderMenu.add_command(label='Info', command=lambda: insrtConsole(f'Merges Excel files in a folder into a singular Excel file\n\n', console))

c2xButton.grid(row=0, column=0, padx=2, pady=2)
x2cButton.grid(row=1, column=0, padx=2, pady=2)
asunderButton.grid(row=0, column=1, padx=2, pady=2)
unasunderButton.grid(row=1, column=1, padx=2, pady=2)

ttk.Separator(root, bootstyle='success', orient='horizontal').pack(fill='x')

consoleFrame = ttk.LabelFrame(root, text=' Console ', bootstyle='success')
consoleFrame.pack(pady=(0,10), padx=10, fill='both', expand=True)

console = ttk.Text(consoleFrame, wrap=tk.WORD, width=1, height=1, font='consolas')
console.pack(pady=0, padx=10, fill=tk.BOTH, expand=True)
console.bind("<Key>", lambda e: 'break')
console.insert(tk.END, 'Welcome to slothyTools\n\n')
ttk.Button(consoleFrame, text='Clear Console', command=clrConsole, bootstyle='success', width=20).pack(pady=(0,10), padx=10, fill='x', expand=False)

root.mainloop()